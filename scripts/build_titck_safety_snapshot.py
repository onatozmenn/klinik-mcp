#!/usr/bin/env python
"""Build data/titck_safety.json from two official TİTCK drug-safety lists.

Sources (titck.gov.tr):
  * Ek İzlemeye Tabi İlaçlar Listesi  (dinamikmodul/57) -> additional monitoring
  * Ruhsat İptal Listesi              (dinamikmodul/76) -> authorization cancellations

Each list is published as an .xlsx attachment. Download the latest of each and:

    python scripts/build_titck_safety_snapshot.py \
        --monitoring ekizleme.xlsx --cancellations ruhsatiptal.xlsx \
        --monitoring-version 19.12.2025 --cancellations-version 19.06.2026

Either file is optional — pass whichever you have; the other keeps its previous
records only if you also merge manually (this script writes a fresh snapshot).
Requires openpyxl (``pip install -e .[dev]``).
"""
from __future__ import annotations

import argparse
import datetime as dt
import json
import sys
import unicodedata
from pathlib import Path

try:
    from openpyxl import load_workbook
except ModuleNotFoundError:
    sys.exit("openpyxl gerekli: pip install openpyxl")

OUT_PATH = (
    Path(__file__).resolve().parents[1]
    / "src"
    / "health_mcp"
    / "data"
    / "titck_safety.json"
)


def norm(value) -> str:
    text = "" if value is None else str(value)
    text = text.upper().translate(str.maketrans("ÇĞİÖŞÜ", "CGIOSU"))
    text = unicodedata.normalize("NFKD", text)
    text = "".join(c for c in text if not unicodedata.combining(c))
    return " ".join(text.split())


def _text(value):
    return " ".join(str(value).split()) if value not in (None, "") else None


def to_date(value):
    if isinstance(value, (dt.datetime, dt.date)):
        return value.isoformat()[:10]
    text = str(value).strip() if value else ""
    return text[:10] or None


def find_header_row(ws, predicate, max_scan: int = 8):
    for i, row in enumerate(
        ws.iter_rows(min_row=1, max_row=max_scan, values_only=True), start=1
    ):
        cells = [norm(c) for c in row]
        if predicate(cells):
            return i, row
    return None, None


def parse_monitoring(path: str) -> list[dict]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = next(
        (
            s
            for s in workbook.worksheets
            if find_header_row(
                s, lambda c: any("ILAC ADI" in x for x in c) and any("ETKIN MADDE" in x for x in c)
            )[0]
        ),
        workbook.active,
    )
    header_idx, header_row = find_header_row(
        worksheet,
        lambda c: any("ILAC ADI" in x for x in c) and any("ETKIN MADDE" in x for x in c),
    )
    if not header_row:
        sys.exit("Ek izleme: başlık satırı bulunamadı.")
    cols: dict[str, int] = {}
    for idx, cell in enumerate(header_row):
        h = norm(cell)
        if "name" not in cols and "ILAC ADI" in h:
            cols["name"] = idx
        elif "active" not in cols and "ETKIN MADDE" in h:
            cols["active"] = idx
        elif "date" not in cols and "LISTEYE" in h:
            cols["date"] = idx

    records: list[dict] = []
    seen: set[tuple] = set()
    for row in worksheet.iter_rows(min_row=header_idx + 1, values_only=True):
        def get(field):
            idx = cols.get(field)
            return row[idx] if idx is not None and idx < len(row) else None

        name = _text(get("name"))
        if not name:
            continue
        key = (norm(name), norm(get("active")))
        if key in seen:
            continue
        seen.add(key)
        records.append(
            {
                "name": name,
                "active": _text(get("active")),
                "date": to_date(get("date")),
            }
        )
    return records


def parse_cancellations(path: str) -> list[dict]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    records: list[dict] = []
    seen: set[tuple] = set()
    for worksheet in workbook.worksheets:
        header_idx, header_row = find_header_row(
            worksheet,
            lambda c: any("ILAC ADI" in x for x in c)
            and any(("BARKOD" in x or "RUHSAT" in x) for x in c),
        )
        if not header_row:
            continue
        cols: dict[str, int] = {}
        for idx, cell in enumerate(header_row):
            h = norm(cell)
            if "name" not in cols and "ILAC ADI" in h:
                cols["name"] = idx
            elif "holder" not in cols and "RUHSAT SAHIBI" in h:
                cols["holder"] = idx
            elif h == "BARKOD":
                cols["barcode"] = idx
            elif "MAKAM OLURU" in h and ("NO" in h or "cancel_date" not in cols):
                cols["cancel_date"] = idx

        for row in worksheet.iter_rows(min_row=header_idx + 1, values_only=True):
            def get(field):
                idx = cols.get(field)
                return row[idx] if idx is not None and idx < len(row) else None

            name = _text(get("name"))
            if not name:
                continue
            cancel_date = _text(get("cancel_date"))
            key = (norm(name), norm(cancel_date))
            if key in seen:
                continue
            seen.add(key)
            barcode = get("barcode")
            records.append(
                {
                    "name": name,
                    "holder": _text(get("holder")),
                    "barcode": str(barcode).strip().split(".")[0] if barcode else None,
                    "cancel_date": cancel_date,
                    "sheet": worksheet.title,
                }
            )
    return records


def main() -> None:
    parser = argparse.ArgumentParser(description="TİTCK güvenlik listeleri → JSON snapshot")
    parser.add_argument("--monitoring", help="Ek İzleme .xlsx yolu")
    parser.add_argument("--cancellations", help="Ruhsat İptal .xlsx yolu")
    parser.add_argument("--monitoring-version", default="")
    parser.add_argument("--cancellations-version", default="")
    args = parser.parse_args()

    if not args.monitoring and not args.cancellations:
        sys.exit("En az bir liste verin (--monitoring ve/veya --cancellations).")

    monitoring = parse_monitoring(args.monitoring) if args.monitoring else []
    cancellations = parse_cancellations(args.cancellations) if args.cancellations else []

    payload = {
        "meta": {
            "monitoring": {
                "source": "TİTCK Ek İzlemeye Tabi İlaçlar Listesi (dinamikmodul/57)",
                "version": args.monitoring_version or "bilinmiyor",
                "count": len(monitoring),
            },
            "cancellations": {
                "source": "TİTCK Ruhsat İptal Listesi (dinamikmodul/76)",
                "version": args.cancellations_version or "bilinmiyor",
                "count": len(cancellations),
            },
            "sample": False,
        },
        "monitoring": monitoring,
        "cancellations": cancellations,
    }
    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    with OUT_PATH.open("w", encoding="utf-8") as handle:
        json.dump(payload, handle, ensure_ascii=False)
    print(
        f"{len(monitoring)} ek izleme + {len(cancellations)} ruhsat iptali yazıldı → {OUT_PATH}"
    )


if __name__ == "__main__":
    main()
