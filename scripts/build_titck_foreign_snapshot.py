#!/usr/bin/env python
"""Build data/titck_foreign.json from TİTCK's "Yurt Dışı Etkin Madde Listesi".

Source: titck.gov.tr → dinamikmodul/126 "Yurt Dışı Etkin Madde Listesi" → latest
.xlsx. This is the list of active substances that may be supplied from abroad
(via TEB), with ATC, pharmaceutical form, prescription type, whether import is
allowed without TİTCK's written approval, and any ICD-10 codes.

Usage:
    python scripts/build_titck_foreign_snapshot.py path/to/yurtdisi.xlsx --version 2026-06-25

Requires openpyxl (``pip install -e .[dev]``).
"""
from __future__ import annotations

import argparse
import json
import sys
import unicodedata
from pathlib import Path

try:
    from openpyxl import load_workbook
except ModuleNotFoundError:
    sys.exit("openpyxl gerekli: pip install openpyxl")

OUT_PATH = (
    Path(__file__).resolve().parents[1]
    / "src"
    / "health_mcp"
    / "data"
    / "titck_foreign.json"
)


def norm(value) -> str:
    text = "" if value is None else str(value)
    text = text.upper().translate(str.maketrans("ÇĞİÖŞÜ", "CGIOSU"))
    text = unicodedata.normalize("NFKD", text)
    text = "".join(c for c in text if not unicodedata.combining(c))
    return " ".join(text.split())


def _text(value):
    return " ".join(str(value).split()) if value not in (None, "") else None


def find_header_row(ws, max_scan: int = 8):
    for i, row in enumerate(
        ws.iter_rows(min_row=1, max_row=max_scan, values_only=True), start=1
    ):
        cells = [norm(c) for c in row]
        if any("ETKIN MADDE" in c for c in cells) and any("ATC" in c for c in cells):
            return i, row
    return None, None


def map_columns(header_row) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for idx, cell in enumerate(header_row):
        h = norm(cell)
        if "code" not in mapping and "ETKIN MADDE KODU" in h:
            mapping["code"] = idx
        elif "active" not in mapping and h == "ETKIN MADDE":
            mapping["active"] = idx
        elif "form" not in mapping and "FORM" in h:
            mapping["form"] = idx
        elif "atc_code" not in mapping and h == "ATC KODU":
            mapping["atc_code"] = idx
        elif "atc_name" not in mapping and "ATC ADI" in h:
            mapping["atc_name"] = idx
        elif "prescription_type" not in mapping and "RECETE TURU" in h:
            mapping["prescription_type"] = idx
        elif "import_flag" not in mapping and "ONAYI OLMADAN" in h:
            mapping["import_flag"] = idx
        elif "icd10" not in mapping and "ICD" in h:
            mapping["icd10"] = idx
    return mapping


def main() -> None:
    parser = argparse.ArgumentParser(description="TİTCK Yurt Dışı Etkin Madde → JSON")
    parser.add_argument("excel", help="Path to the Yurt Dışı Etkin Madde .xlsx")
    parser.add_argument("--version", default="", help="Version label, e.g. 2026-06-25")
    args = parser.parse_args()

    workbook = load_workbook(args.excel, read_only=True, data_only=True)
    worksheet = next(
        (s for s in workbook.worksheets if find_header_row(s)[0]), workbook.active
    )
    header_idx, header_row = find_header_row(worksheet)
    if not header_row:
        sys.exit("Başlık satırı bulunamadı.")
    mapping = map_columns(header_row)
    print("Eşlenen sütunlar:", mapping)
    if not {"active", "atc_code"} <= mapping.keys():
        sys.exit("Zorunlu sütunlar (active, atc_code) eşlenemedi.")

    substances = []
    seen: set[str] = set()
    for row in worksheet.iter_rows(min_row=header_idx + 1, values_only=True):

        def get(field):
            idx = mapping.get(field)
            return row[idx] if idx is not None and idx < len(row) else None

        active = _text(get("active"))
        if not active:
            continue
        code = _text(get("code"))
        key = code or active
        if key in seen:
            continue
        seen.add(key)
        flag = norm(get("import_flag"))
        substances.append(
            {
                "code": code,
                "active": active,
                "form": _text(get("form")),
                "atc_code": _text(get("atc_code")),
                "atc_name": _text(get("atc_name")),
                "prescription_type": _text(get("prescription_type")),
                "import_without_approval": flag in {"1", "1.0", "EVET", "VAR", "X"},
                "icd10": _text(get("icd10")),
            }
        )

    if not substances:
        sys.exit("Hiç kayıt bulunamadı.")

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    payload = {
        "meta": {
            "source": "TİTCK Yurt Dışı Etkin Madde Listesi (dinamikmodul/126)",
            "version": args.version or "bilinmiyor",
            "count": len(substances),
        },
        "substances": substances,
    }
    with OUT_PATH.open("w", encoding="utf-8") as handle:
        json.dump(payload, handle, ensure_ascii=False)
    print(f"{len(substances)} kayıt yazıldı → {OUT_PATH}")


if __name__ == "__main__":
    main()
