#!/usr/bin/env python
"""Build data/titck_drugs.json from the TİTCK SKRS E-Reçete drug list Excel.

Source: titck.gov.tr → dinamikmodul/43 "SKRS E-Reçete İlaç ve Diğer Farmasötik
Ürünler Listesi" → latest .xlsx (sheet "AKTİF ÜRÜNLER LİSTESİ").

Usage:
    python scripts/build_titck_snapshot.py path/to/skrs.xlsx --version 2026-06-23

Requires openpyxl (``pip install openpyxl`` or ``pip install -e .[dev]``).
"""
from __future__ import annotations

import argparse
import json
import sys
import unicodedata
from pathlib import Path

try:
    from openpyxl import load_workbook
except ModuleNotFoundError:
    sys.exit("openpyxl gerekli: pip install openpyxl")

OUT_PATH = (
    Path(__file__).resolve().parents[1]
    / "src"
    / "health_mcp"
    / "data"
    / "titck_drugs.json"
)

COLUMN_RULES = [
    ("name", lambda h: "ILAC ADI" in h),
    ("barcode", lambda h: "BARKOD" in h),
    ("atc_code", lambda h: h == "ATC KODU"),
    ("atc_name", lambda h: "ATC ADI" in h),
    ("company", lambda h: "FIRMA" in h),
    ("prescription_type", lambda h: "RECETE TURU" in h),
    ("essential", lambda h: h.startswith("TEMEL ILAC")),
]


def norm(value) -> str:
    text = "" if value is None else str(value)
    text = text.upper().translate(str.maketrans("ÇĞİÖŞÜ", "CGIOSU"))
    text = unicodedata.normalize("NFKD", text)
    text = "".join(c for c in text if not unicodedata.combining(c))
    return " ".join(text.split())


def find_header_row(ws, max_scan: int = 10):
    for i, row in enumerate(
        ws.iter_rows(min_row=1, max_row=max_scan, values_only=True), start=1
    ):
        cells = [norm(c) for c in row]
        if any("ILAC ADI" in c for c in cells) and any("BARKOD" in c for c in cells):
            return i, row
    return None, None


def map_columns(header_row) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for idx, cell in enumerate(header_row):
        header = norm(cell)
        for field, rule in COLUMN_RULES:
            if field not in mapping and rule(header):
                mapping[field] = idx
                break
    return mapping


def _text(value):
    return " ".join(str(value).split()) if value else None


def main() -> None:
    parser = argparse.ArgumentParser(description="TİTCK SKRS → JSON snapshot")
    parser.add_argument("excel", help="Path to the SKRS .xlsx file")
    parser.add_argument("--version", default="", help="Version label, e.g. 2026-06-23")
    args = parser.parse_args()

    workbook = load_workbook(args.excel, read_only=True, data_only=True)
    worksheet = next(
        (s for s in workbook.worksheets if "AKTIF" in norm(s.title)), workbook.active
    )
    header_idx, header_row = find_header_row(worksheet)
    if not header_row:
        sys.exit("Başlık satırı bulunamadı.")
    mapping = map_columns(header_row)
    print("Eşlenen sütunlar:", mapping)
    if not {"name", "barcode"} <= mapping.keys():
        sys.exit("Zorunlu sütunlar (name, barcode) eşlenemedi.")

    drugs = []
    seen: set[str] = set()
    for row in worksheet.iter_rows(min_row=header_idx + 1, values_only=True):

        def get(field):
            idx = mapping.get(field)
            return row[idx] if idx is not None and idx < len(row) else None

        name, barcode = get("name"), get("barcode")
        if not name or not barcode:
            continue
        barcode = str(barcode).strip()
        if barcode in seen:
            continue
        seen.add(barcode)
        essential = get("essential")
        drugs.append(
            {
                "name": _text(name),
                "barcode": barcode,
                "atc_code": _text(get("atc_code")),
                "atc_name": _text(get("atc_name")),
                "company": _text(get("company")),
                "prescription_type": _text(get("prescription_type")),
                "essential": str(essential).strip() in {"1", "1.0", "EVET", "VAR"}
                if essential is not None
                else False,
            }
        )

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    payload = {
        "meta": {
            "source": "TİTCK SKRS E-Reçete İlaç Listesi (Aktif Ürünler)",
            "version": args.version or "bilinmiyor",
            "count": len(drugs),
        },
        "drugs": drugs,
    }
    with OUT_PATH.open("w", encoding="utf-8") as handle:
        json.dump(payload, handle, ensure_ascii=False)
    print(f"{len(drugs)} kayıt yazıldı → {OUT_PATH}")


if __name__ == "__main__":
    main()
