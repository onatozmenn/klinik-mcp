#!/usr/bin/env python
"""Build the bundled SGK EK-4/A JSON snapshot from an official SGK Excel file.

Usage:
    python scripts/build_sgk_snapshot.py path/to/EK-4A.xlsx --version "2026/23"

Requires openpyxl (``pip install openpyxl`` or ``pip install -e .[dev]``).

Where to get the source file:
    * Periodic amendment files ("4A Eklenenler/Düzenlenenler") are published on
      sgk.gov.tr → Duyurular → "Bedeli Ödenecek İlaçlar Listesinde Yapılan
      Düzenlemeler".
    * The full consolidated EK-4/A is distributed via MEDULA / the Türk
      Eczacıları Birliği (teb.org.tr). Drop that .xlsx in and re-run this script.

The parser handles the real EK-4/A layout: title rows on top, the header on a
later row, and multiple worksheets ("4A ...").
"""
from __future__ import annotations

import argparse
import datetime as dt
import json
import sys
import unicodedata
from pathlib import Path

try:
    from openpyxl import load_workbook
except ModuleNotFoundError:
    sys.exit("openpyxl gerekli: pip install openpyxl")

OUT_PATH = (
    Path(__file__).resolve().parents[1]
    / "src"
    / "health_mcp"
    / "data"
    / "sgk_ek4a.json"
)

# field -> predicate over the normalized header text
COLUMN_RULES = [
    ("kamu_no", lambda h: h == "KAMU NO"),
    ("barcode", lambda h: "GUNCEL BARKOD" in h),
    ("name", lambda h: "ILAC ADI" in h),
    ("equivalent_group", lambda h: "ESDEGER" in h and ("GRUP" in h or "GRUB" in h)),
    ("entry_date", lambda h: "LISTEYE GIRIS" in h),
]


def norm(value) -> str:
    text = "" if value is None else str(value)
    text = text.upper().translate(str.maketrans("ÇĞİÖŞÜ", "CGIOSU"))
    text = unicodedata.normalize("NFKD", text)
    text = "".join(c for c in text if not unicodedata.combining(c))
    return " ".join(text.split())


def find_header_row(ws, max_scan: int = 15):
    for i, row in enumerate(
        ws.iter_rows(min_row=1, max_row=max_scan, values_only=True), start=1
    ):
        cells = [norm(c) for c in row]
        if any("ILAC ADI" in c for c in cells) and any("BARKOD" in c for c in cells):
            return i, row
    return None, None


def map_columns(header_row) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for idx, cell in enumerate(header_row):
        header = norm(cell)
        for field, rule in COLUMN_RULES:
            if field not in mapping and rule(header):
                mapping[field] = idx
                break
    return mapping


def to_float(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text:
        return None
    if "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".")
    elif "," in text:
        text = text.replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return None


def to_date(value):
    if isinstance(value, (dt.datetime, dt.date)):
        return value.isoformat()[:10]
    text = str(value).strip() if value else ""
    return text[:10] or None


def main() -> None:
    parser = argparse.ArgumentParser(description="SGK EK-4/A → JSON snapshot")
    parser.add_argument("excel", help="Path to the EK-4/A .xlsx file")
    parser.add_argument("--version", default="", help="Version label, e.g. 2026/23")
    args = parser.parse_args()

    workbook = load_workbook(args.excel, read_only=True, data_only=True)
    drugs: list[dict] = []
    seen: set[str] = set()
    sheets_used: list[str] = []

    for ws in workbook.worksheets:
        if "4A" not in norm(ws.title):
            continue
        header_idx, header_row = find_header_row(ws)
        if not header_row:
            continue
        mapping = map_columns(header_row)
        if not {"barcode", "name"} <= mapping.keys():
            continue
        sheets_used.append(ws.title)

        for row in ws.iter_rows(min_row=header_idx + 1, values_only=True):

            def get(field):
                idx = mapping.get(field)
                return row[idx] if idx is not None and idx < len(row) else None

            barcode, name = get("barcode"), get("name")
            if not barcode or not name:
                continue
            barcode = str(barcode).strip()
            if barcode in seen:
                continue
            seen.add(barcode)
            drugs.append(
                {
                    "kamu_no": str(get("kamu_no")).strip() if get("kamu_no") else None,
                    "barcode": barcode,
                    "name": " ".join(str(name).split()),
                    "equivalent_group": str(get("equivalent_group")).strip()
                    if get("equivalent_group")
                    else None,
                    "entry_date": to_date(get("entry_date")),
                    "reimbursed": True,
                }
            )

    if not drugs:
        sys.exit("Hiç kayıt bulunamadı (sayfa/başlık eşleşmedi).")

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    payload = {
        "meta": {
            "source": "SGK EK-4/A Bedeli Ödenecek İlaçlar Listesi",
            "version": args.version or "bilinmiyor",
            "sample": False,
            "sheets": sheets_used,
            "count": len(drugs),
        },
        "drugs": drugs,
    }
    with OUT_PATH.open("w", encoding="utf-8") as handle:
        json.dump(payload, handle, ensure_ascii=False)
    print(f"{len(drugs)} kayıt yazıldı (sayfalar: {sheets_used}) → {OUT_PATH}")


if __name__ == "__main__":
    main()
