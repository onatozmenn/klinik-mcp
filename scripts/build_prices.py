#!/usr/bin/env python
"""Build data/prices.json from a barcode -> TL price export (CSV or Excel).

The source is any commercial / pharmacy-automation export (e.g. Rxmediapharma,
eczane otomasyonu) that has at least a barcode column and a retail (perakende)
and/or depot (depocu) price column. Column names are auto-detected.

Usage:
    python scripts/build_prices.py path/to/prices.csv  --source "Rxmediapharma" --version 2026-06
    python scripts/build_prices.py path/to/prices.xlsx --source "Eczane otomasyonu"

Requires openpyxl only for .xlsx inputs (pip install -e .[dev]).
"""
from __future__ import annotations

import argparse
import csv
import json
import sys
import unicodedata
from pathlib import Path

OUT_PATH = (
    Path(__file__).resolve().parents[1]
    / "src"
    / "health_mcp"
    / "data"
    / "prices.json"
)


def norm(value) -> str:
    text = "" if value is None else str(value)
    text = text.upper().translate(str.maketrans("ÇĞİÖŞÜ", "CGIOSU"))
    text = unicodedata.normalize("NFKD", text)
    text = "".join(c for c in text if not unicodedata.combining(c))
    return " ".join(text.split())


def detect_columns(headers: list[str]) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for idx, header in enumerate(headers):
        h = norm(header)
        if "barcode" not in mapping and "BARKOD" in h:
            mapping["barcode"] = idx
        elif "retail" not in mapping and (
            "PERAKENDE" in h or "PSF" in h or "KDV" in h
        ):
            mapping["retail"] = idx
        elif "depot" not in mapping and (
            "DEPOCU" in h or "DSF" in h or ("DEPO" in h and "FIYAT" in h)
        ):
            mapping["depot"] = idx
    return mapping


def to_float(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text:
        return None
    if "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".")
    elif "," in text:
        text = text.replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return None


def read_rows(path: Path) -> tuple[list[str], list[list]]:
    if path.suffix.lower() in {".xlsx", ".xlsm"}:
        try:
            from openpyxl import load_workbook
        except ModuleNotFoundError:
            sys.exit("xlsx için openpyxl gerekli: pip install openpyxl")
        ws = load_workbook(path, read_only=True, data_only=True).active
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
    else:
        with path.open(encoding="utf-8-sig", newline="") as handle:
            sample = handle.read(4096)
            handle.seek(0)
            delimiter = ";" if sample.count(";") > sample.count(",") else ","
            rows = [row for row in csv.reader(handle, delimiter=delimiter)]
    # find the header row (first row containing BARKOD)
    for i, row in enumerate(rows[:15]):
        if any("BARKOD" in norm(c) for c in row):
            return [str(c) for c in row], rows[i + 1 :]
    return ([str(c) for c in rows[0]], rows[1:]) if rows else ([], [])


def main() -> None:
    parser = argparse.ArgumentParser(description="barcode->price -> prices.json")
    parser.add_argument("path", help="CSV or Excel file with barcode + price columns")
    parser.add_argument("--source", default="ticari kaynak")
    parser.add_argument("--version", default="")
    args = parser.parse_args()

    headers, rows = read_rows(Path(args.path))
    mapping = detect_columns(headers)
    print("Eşlenen sütunlar:", mapping)
    if "barcode" not in mapping or not ({"retail", "depot"} & mapping.keys()):
        sys.exit("Barkod ve en az bir fiyat (perakende/depocu) sütunu gerekli.")

    prices: dict[str, dict] = {}
    for row in rows:
        def get(field):
            idx = mapping.get(field)
            return row[idx] if idx is not None and idx < len(row) else None

        barcode = get("barcode")
        if not barcode:
            continue
        barcode = str(barcode).strip().split(".")[0]  # strip any .0 from numeric cells
        entry = {}
        if "retail" in mapping:
            entry["retail"] = to_float(get("retail"))
        if "depot" in mapping:
            entry["depot"] = to_float(get("depot"))
        if any(v is not None for v in entry.values()):
            prices[barcode] = entry

    OUT_PATH.write_text(
        json.dumps(
            {
                "meta": {
                    "source": args.source,
                    "version": args.version or "bilinmiyor",
                    "currency": "TRY",
                    "sample": False,
                    "count": len(prices),
                },
                "prices": prices,
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    print(f"{len(prices)} fiyat yazıldı → {OUT_PATH}")


if __name__ == "__main__":
    main()
